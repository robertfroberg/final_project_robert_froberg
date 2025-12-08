import sys
import os
import re
import pandas as pd
import openpyxl
from difflib import get_close_matches
import requests

from config import DATA_DIR, MAGIC_ITEM_URL


# download magic item workbook from url into data folder
def download_magic_items_xlsx(local_path):
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    print(f"Downloading magic item workbook from URL to: {local_path}")
    with requests.get(MAGIC_ITEM_URL, stream=True) as r:
        r.raise_for_status()
        with open(local_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
    return local_path


# load default magic item workbook from data folder, downloading if needed
def load_default_magic_items(xlsx_filename="MagicItemList.xlsx"):
    file_path = os.path.join(DATA_DIR, xlsx_filename)

    if not os.path.exists(file_path):
        try:
            download_magic_items_xlsx(file_path)
        except Exception as e:
            print(f" Error downloading magic item file: {e}")
            return None

    try:
        df = load_magic_items(file_path)
        return df
    except Exception as e:
        print("error loading magic items:", e)
        return None


# fuzzy match pc name against owner column
def fuzzy_find_owner(target_name, owner_list, cutoff=0.6):
    if not target_name:
        return None

    norm_target = target_name.strip().lower()
    norm_list = [o.strip().lower() for o in owner_list]

    matches = get_close_matches(norm_target, norm_list, n=1, cutoff=cutoff)
    if not matches:
        return None

    idx = norm_list.index(matches[0])
    return owner_list[idx]


# get items for pc name using fuzzy owner match
def get_items_for_pc_name(pc_name, dataframe, cutoff=0.6):
    if dataframe is None or dataframe.empty:
        return None, dataframe

    if not pc_name:
        return None, dataframe

    owners = dataframe["owner"].dropna().unique()
    matched_owner = fuzzy_find_owner(pc_name, owners, cutoff=cutoff)

    if matched_owner is None:
        return None, dataframe

    items = dataframe[dataframe["owner"] == matched_owner]
    return matched_owner, items


# pretty print pc magic items and summary
def print_items_for_pc_name(pc_name, dataframe, cutoff=0.6):
    print_header = lambda title: print("\n" + "=" * 60 + f"\n{title}\n" + "=" * 60)

    print_header("Magic Item List")

    if dataframe is None or dataframe.empty:
        print("No magic item data loaded.")
        return

    if not pc_name:
        print("PC name not found on character object.")
        return

    matched_owner, items = get_items_for_pc_name(pc_name, dataframe, cutoff=cutoff)

    if matched_owner is None:
        owners = dataframe["owner"].dropna().unique()
        print(f"No owner found matching '{pc_name}'.")
        print("Available owners:", ", ".join(owners))
        return

    if matched_owner != pc_name:
        print(f"Matched '{pc_name}' → '{matched_owner}'")

    print(f"Owner: {matched_owner}")

    if items.empty:
        print("No magic items found for this character.")
        return

    total = len(items)
    carried_total = items["carried"].sum()
    rarity_counts = items["rarity"].value_counts()

    print(f"Total items: {total}")
    print(f"Carried items: {carried_total}")
    print("Items by rarity:")
    for rarity, count in rarity_counts.items():
        print(f"  {rarity}: {count}")

    print("\nItem List:")
    for _, row in items.iterrows():
        tag = "*" if row["carried"] else ""
        prop = f" ({row['spec_prop']})" if row["spec_prop"] else ""
        print(f"  {row['item_name']}{prop} {tag}")


# load workbook and target sheet
def load_magic_items(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook["Items_By_Rarity"]

    df = pd.DataFrame(
        columns=[
            "owner",
            "item_name",
            "spec_prop",
            "rarity",
            "carried",
            "trade_will",
            "reward",
            "certed",
        ]
    )

    # loop owner columns
    for col_index in range(2, sheet.max_column + 1):
        owner = sheet.cell(row=3, column=col_index).value

        # loop item rows
        for row_index in range(4, sheet.max_row + 1):
            item_value = sheet.cell(row=row_index, column=col_index).value
            spec_prop = ""

            if pd.isna(item_value):
                continue

            text = str(item_value)
            match = re.search(r"\((.*?)\)", text)
            if match:
                spec_prop = match.group(1)
                item_name = text.replace(f"({spec_prop})", "").strip()
            else:
                item_name = text

            font = sheet.cell(row=row_index, column=col_index).font
            rgb = getattr(font.color, "rgb", None)

            if rgb == "FFFFC000":
                rarity = "Unique"
            elif rgb == "FF7030A0":
                rarity = "Legendary"
            elif rgb == "FF0070C0":
                rarity = "Very Rare"
            elif rgb == "FF00B050":
                rarity = "Rare"
            elif rgb == "FFFF0000":
                rarity = "Uncommon"
            else:
                rarity = "Common"

            fill = sheet.cell(row=row_index, column=col_index).fill.start_color.index
            if fill == 2:
                trade_will = "Untradeable"
            elif fill == 5:
                trade_will = "Red"
            elif fill == 7:
                trade_will = "Amber"
            elif fill == 9:
                trade_will = "Green"
            else:
                trade_will = "FAILED"
                print(f"[WARN] unexpected fill index ({fill}) at row {row_index}, col {col_index}")

            borders = sheet.cell(row=row_index, column=col_index).border
            carried = (
                borders.top.style == "thin"
                and borders.bottom.style == "thin"
                and borders.left.style == "thin"
                and borders.right.style == "thin"
            )

            reward = font.italic
            certed = font.bold

            df.loc[len(df)] = [
                owner,
                item_name,
                spec_prop,
                rarity,
                carried,
                trade_will,
                reward,
                certed,
            ]

    base_name, _ = os.path.splitext(file_path)
    csv_path = base_name + "_items.csv"
    df.to_csv(csv_path, index=False)
    print("dataframe saved to:", csv_path)

    return df


# list of stat book names
def book_count(dataframe):
    book_names = [
        "Manual of Gainful Exercise",
        "Manual of Quickness of Action",
        "Manual of Bodily Health",
        "Tome of Clear Thought",
        "Tome of Understanding",
        "Tome of Leadership and Influence",
    ]

    print("\n=== Book Count (Tradeable Only) ===")
    results = {}

    for name in book_names:
        count = dataframe[
            (dataframe["item_name"].str.contains(name, na=False))
            & (dataframe["trade_will"] != "Untradeable")
        ].shape[0]

        results[name] = count
        print(f"{name}: {count}")

    return results


# pull unique owner names
def list_owners_and_items(dataframe):
    owners = dataframe["owner"].dropna().unique()

    if len(owners) == 0:
        print("No owners found.")
        return

    print("\n=== Owners ===")
    for i, owner in enumerate(owners, 1):
        print(f"{i}. {owner}")

    while True:
        choice = input("\nSelect a character (name or number): ").strip()

        if choice.isdigit():
            idx = int(choice)
            if 1 <= idx <= len(owners):
                owner = owners[idx - 1]
                break
        elif choice in owners:
            owner = choice
            break

        print("Invalid selection. Try again.")

    items = dataframe[dataframe["owner"] == owner]
    total = len(items)
    carried_total = items["carried"].sum()
    rarity_counts = items["rarity"].value_counts()

    print(f"\n=== Items for {owner} ===")
    print(f"Total items: {total}")
    print(f"Carried items: {carried_total}")
    print("Items by rarity:")
    for r, c in rarity_counts.items():
        print(f"  {r}: {c}")

    print("\nItem List:")
    for _, row in items.iterrows():
        tag = "*" if row["carried"] else ""
        prop = f" ({row['spec_prop']})" if row["spec_prop"] else ""
        print(f"  {row['item_name']}{prop} {tag}")


# generate alphabetized trade list without duplicates
def generate_trade_list(dataframe):
    rarity_map = {
        "common": "Common",
        "uncommon": "Uncommon",
        "rare": "Rare",
        "very rare": "Very Rare",
        "legendary": "Legendary",
    }

    trade_map = {
        "red": ["Red", "Amber", "Green"],
        "amber": ["Amber", "Green"],
        "green": ["Green"],
    }

    def ask(prompt, options):
        while True:
            val = input(prompt).strip().lower()
            if val.isdigit():
                num = int(val)
                if 1 <= num <= len(options):
                    return options[num - 1]
            if val in options:
                return val
            print("Invalid option. Try again.")

    print("\n=== Trade List Generator ===")

    rarity_opts = list(rarity_map.keys())
    rarity_sel = ask(f"Select rarity ({', '.join(rarity_opts)}): ", rarity_opts)
    rarity_label = rarity_map[rarity_sel]

    trade_opts = list(trade_map.keys())
    trade_sel = ask(f"Select trade tier ({', '.join(trade_opts)}): ", trade_opts)
    allowed_tiers = trade_map[trade_sel]

    subset = dataframe[
        (dataframe["rarity"] == rarity_label)
        and (dataframe["trade_will"].isin(allowed_tiers))
    ]

    names = set()
    for _, row in subset.iterrows():
        prop = f" ({row['spec_prop']})" if row["spec_prop"] else ""
        names.add(f"{row['item_name']}{prop}")

    print()
    if names:
        print(f"{rarity_label} Trade Items ({', '.join(allowed_tiers)}):")
        print(", ".join(sorted(names)))
        print()
    else:
        print(f"No {rarity_label} items available for {', '.join(allowed_tiers)}.\n")


def generate_rarity_summary_table(dataframe):
    total = dataframe["rarity"].value_counts().sort_index()
    carried = dataframe[dataframe["carried"]]["rarity"].value_counts().sort_index()
    untrade = (
        dataframe[dataframe["trade_will"] == "Untradeable"]["rarity"]
        .value_counts()
        .sort_index()
    )

    table = pd.DataFrame(
        {
            "Total": total,
            "Red": dataframe[dataframe["trade_will"] == "Red"]["rarity"].value_counts(),
            "Amber": dataframe[dataframe["trade_will"] == "Amber"]["rarity"].value_counts(),
            "Green": dataframe[dataframe["trade_will"] == "Green"]["rarity"].value_counts(),
            "Untradeable": untrade,
            "Carried": carried,
        }
    ).fillna(0).astype(int)

    table.loc["Sum"] = [
        table["Total"].sum(),
        table["Red"].sum(),
        table["Amber"].sum(),
        table["Green"].sum(),
        table["Untradeable"].sum(),
        table["Carried"].sum(),
    ]

    print("\n=== Rarity Summary Table ===")
    print(table)
    print()
    return table


# interactive tool entry point
def main():
    if len(sys.argv) < 2:
        print("usage: python magic_items.py <MagicItemList.xlsx>")
        sys.exit(1)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "..", "data")

    xlsx_filename = os.path.basename(sys.argv[1])
    file_path = os.path.join(data_dir, xlsx_filename)

    if not os.path.exists(file_path):
        print(f"error: file not found in /data: {file_path}")
        sys.exit(1)

    try:
        df = load_magic_items(file_path)
    except Exception as e:
        print("error loading file:", e)
        sys.exit(1)

    while True:
        print("\n=== Magic Item Tool ===")
        print("1. Book count")
        print("2. List owners and items")
        print("3. Generate a trade list")
        print("4. Rarity summary table")
        print("5. Exit")

        choice = input("Select an option: ").strip()

        if choice == "1":
            book_count(df)
        elif choice == "2":
            list_owners_and_items(df)
        elif choice == "3":
            generate_trade_list(df)
        elif choice == "4":
            generate_rarity_summary_table(df)
        elif choice == "5":
            print("Exiting.")
            break
        else:
            print("Invalid selection.\n")


if __name__ == "__main__":
    main()